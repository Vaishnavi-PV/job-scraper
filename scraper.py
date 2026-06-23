import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from urllib.parse import urljoin
import os

print("🚀 Script started...")

# URLs
base_url = "https://realpython.github.io"
url = base_url + "/fake-jobs/"

headers = {
    "User-Agent": "Mozilla/5.0"
}

# Request page
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.text, "html.parser")

jobs = soup.find_all("div", class_="card-content")

job_list = []

# -------- Helper Functions -------- #

def get_experience(title):
    title = title.lower()

    if "senior" in title:
        return "5+ years"
    elif "lead" in title or "manager" in title:
        return "7+ years"
    elif "junior" in title:
        return "1-2 years"
    else:
        return "2-4 years"


def get_skills(title):
    title = title.lower()

    if "python" in title:
        return "Python, Django, Flask, SQL"
    elif "engineer" in title:
        return "Python, Java, SQL, Git"
    elif "developer" in title:
        return "JavaScript, React, HTML, CSS"
    elif "data" in title:
        return "Python, Pandas, Machine Learning, SQL"
    else:
        return "Communication, Problem Solving, Teamwork"


def get_description(title):
    return f"We are looking for a {title} who can contribute by building scalable solutions, collaborating with teams, and delivering high-quality applications."


# -------- Scraping -------- #

for job in jobs:

    title_tag = job.find("h2", class_="title")
    location_tag = job.find("p", class_="location")
    link_tag = job.find("a")

    title = title_tag.text.strip() if title_tag else "N/A"
    location = location_tag.text.strip() if location_tag else "N/A"

    # Create proper full URL
    link = urljoin(url, link_tag["href"]) if link_tag else "N/A"

    experience = get_experience(title)
    skills = get_skills(title)
    salary = "Not Disclosed"
    description = get_description(title)

    job_list.append({
        "JobTitle": title,
        "Location": location,
        "ExperienceRequired": experience,
        "SkillsRequired": skills,
        "Salary": salary,
        "JobURL": link,
        "JobDescriptionSummary": description
    })


# -------- Create DataFrame -------- #

df = pd.DataFrame(job_list)

# -------- Save Excel -------- #

file_name = "Final_Jobs.xlsx"

# Remove old file if exists
if os.path.exists(file_name):
    try:
        os.remove(file_name)
    except:
        print("⚠️ Please close the Excel file and run again.")
        exit()


with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Jobs")


# -------- Format Excel -------- #

wb = load_workbook(file_name)
ws = wb["Jobs"]

for column in ws.columns:

    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:

        if cell.value:

            # Wrap text
            cell.alignment = Alignment(wrap_text=True)

            # Auto width
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[column_letter].width = min(max_length + 5, 50)


# -------- Make JobURL Clickable -------- #

for cell in ws["F"][1:]:      # Column F = JobURL

    if cell.value != "N/A":
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"


# Save workbook
wb.save(file_name)

print("✅ Done!")
print("📂 Excel file created:", file_name)
print("🔗 Job URLs are clickable.")